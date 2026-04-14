from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage
from openpyxl import Workbook, load_workbook
import re
import os

app = Flask(__name__)



LINE_CHANNEL_ACCESS_TOKEN = os.getenv("CHANNEL_ACCESS_TOKEN")
LINE_CHANNEL_SECRET = os.getenv("CHANNEL_SECRET")

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

EXCEL_FILE = '模具紀錄.xlsx'

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["出廠日期", "入廠日期", "名稱", "原因"])
        wb.save(EXCEL_FILE)

def parse_message(text):
    data = {"出廠日期": "", "入廠日期": "", "名稱": "", "原因": ""}
    for key in data:
        match = re.search(f"{key}\\s*(.*)", text)
        if match:
            data[key] = match.group(1).strip()
    return data

def save_to_excel(data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([data["出廠日期"], data["入廠日期"], data["名稱"], data["原因"]])
    wb.save(EXCEL_FILE)

@app.route("/callback", methods=['POST'])
def callback():
    signature = request.headers['X-Line-Signature']
    body = request.get_data(as_text=True)
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    return 'OK'

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    text = event.message.text
    if "模具出廠通知" in text or "模具入廠通知" in text:
        data = parse_message(text)
        save_to_excel(data)
        line_bot_api.reply_message(event.reply_token, TextMessage(text="✅ 已記錄"))

if __name__ == "__main__":
    init_excel()
    app.run(host="0.0.0.0", port=10000)
